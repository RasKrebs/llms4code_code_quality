import os
import pandas as pd
from typing import Union
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class QualityReport:
    def __init__(self) -> None:
        """
        Results class. Takes results from different sources and stores them in a dataframe. 
        Prior to saving the results, the dataframe is transformed and finally the dataframe
        is saved
        """
        # Create a dataframe to store the results
        self.df = pd.DataFrame(columns=["metric", "framework", "model", "value"])
        
    def __call__(self, new_results: Union[pd.DataFrame, list[pd.DataFrame]]):
        """Call the results to append new results to the dataframe

        Args:
            new_results (pd.DataFrame): Pandas dataframe with results with the columns: metric, framework, model, value
        """
        if isinstance(new_results, list):
            self.df = pd.concat([self.df] + new_results, ignore_index=True)
        else:
            self.df = pd.concat([self.df, new_results], ignore_index=True)
    
    def _transform(self):
        # Transform the results dataframe
        df_pivot = self.df.pivot(index=['model', 'prompt'], columns='metric', values='value').reset_index()
        df_pivot = df_pivot.rename_axis(None, axis=1)
        
        # Set the new dataframe
        self.final_df = df_pivot
        
    def save_results(self, 
                     sheet_name: str):
        """Save the results to an Excel file
        
        Args:
            excel_path (str, optional): Path to the Excel file. Defaults to None.
            excel_file (str, optional): Name of the Excel file. Defaults to "results.xlsx".
        """
        
        # Load the existing Excel file
        path = "results.xlsx"
        
        # Load the existing Excel file
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()

        # Create a new sheet
        new_sheet_name = sheet_name
        ws = wb.create_sheet(title=new_sheet_name)

        # Transform the dataframe
        self._transform()
        
        df = self.final_df.copy()
        
        # Write the DataFrame to the new sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the changes
        wb.save(path)